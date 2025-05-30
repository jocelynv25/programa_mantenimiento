import os, cv2
from tkinter import messagebox
import subprocess
import tkinter as tk
import colores
import sys
import csv
import unicodedata

import openpyxl
import os
import tempfile
import pythoncom
import win32com.client as win32
from tkinter import messagebox

def btnGenerar():
    campos = [
        entryNombre.get(),
        entryCodigo.get(),
        entryDepa.get(),
        entryNombreSolicitante.get(),
        entryNumOrden.get(),
        entryFecha.get()
    ]

    # Verifica si todos los campos tienen contenido
    if not all(campo.strip() for campo in campos):
        messagebox.showerror("Error", "Por favor, complete todos los campos.")
        return
    
    try:
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook("orden_de_trabajo.xlsx")
        ws = wb.active

        # Rellenar las celdas con los datos del formulario
        ws["B10"] = campos[0]
        ws["B11"] = campos[1]
        ws["B12"] = campos[2]
        ws["B13"] = campos[3]
        ws["B16"] = campos[4]
        ws["B17"] = campos[5]

        # Guardar en un archivo temporal Excel
        temp_excel = os.path.join(tempfile.gettempdir(), "orden_rellena.xlsx")
        wb.save(temp_excel)
        wb.close()

        # Convertir a PDF usando Excel (solo Windows)
        pythoncom.CoInitialize()  # Requerido si usas en hilo
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb_pdf = excel.Workbooks.Open(temp_excel)
        temp_pdf = os.path.join(tempfile.gettempdir(), f"orden_generada_{campos[4]}.pdf")
        wb_pdf.ExportAsFixedFormat(0, temp_pdf)  # 0 = PDF
        wb_pdf.Close(False)
        excel.Quit()
        del excel

        # Abrir el PDF
        os.startfile(temp_pdf)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al generar el PDF:\n{e}")
    

def volverMenu():
    subprocess.Popen(['python', 'Menu.py'])
    Ventana.destroy()

def volverRecomendaciones():
    subprocess.Popen(['python', 'Recomendaciones.py', maquinaSelected])
    Ventana.destroy()

def on_closing():
    if messagebox.askokcancel("Salir", "¿Seguro que quieres salir?"):
        Ventana.destroy()

def cargar_info_maquina(nombre):

    with open("bd_maquinas.csv", newline='', encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)  # Saltar encabezados

        for fila in reader:
            print(f"fila[0]: {fila[0]}")
            print(f"fila[1]: {fila[1]}")
            if normalizar(fila[0]) == normalizar(nombre):
                entryNombre.insert(0, fila[0])
                entryCodigo.insert(0, fila[1])

                entryNombre.config(state="readonly")
                entryCodigo.config(state="readonly")
                return
        messagebox.showwarning("Error", "Máquina no encontrada.")

def normalizar(texto):
    if not texto:
        return ""
    texto = texto.lower()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return texto.strip()

maquinaSelected= sys.argv[1]

# Crear ventana
Ventana = tk.Tk()
Ventana.title("Generar reporte")

# Obtener dimensiones de la pantalla
screen_width = Ventana.winfo_screenwidth()
screen_height = Ventana.winfo_screenheight()
# Definir dimensiones y calcular posición para centrar la ventana
window_width = 450
window_height = 400
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2
# Establecer la geometría de la ventana
Ventana.geometry(f"{window_width}x{window_height}+{x}+{y}")

Ventana.protocol("WM_DELETE_WINDOW", on_closing)
Ventana.resizable(False, False)

labTitulo = tk.Label(Ventana, text="Generar reporte", font=("Segoe UI", 14, 'bold'))
#labTitulo.grid(column=1,row=1, padx=5, pady=5)
labTitulo.grid(column=0,row=1, columnspan=2, padx=2, pady=5)

labIngrese = tk.Label(Ventana, text="     Ingrese los siguientes datos", font=("Segoe UI", 10))
labIngrese.grid(column=0,row=2, columnspan=2, padx=2, pady=5)

# Etiqueta y entrada para Nombre
labelNombre = tk.Label(Ventana, text="Nombre de la máquina:")
labelNombre.grid(row=3, column=0, padx=10, pady=5)
entryNombre = tk.Entry(Ventana, width=40)
entryNombre.grid(row=3, column=1, padx=10, pady=5)

# Etiqueta y entrada para Dirección
labelCodigo = tk.Label(Ventana, text="Código:")
labelCodigo.grid(row=4, column=0, padx=10, pady=5)
entryCodigo = tk.Entry(Ventana, width=40)
entryCodigo.grid(row=4, column=1, padx=10, pady=5)

labelDepa = tk.Label(Ventana, text="Dpto. que solicita:")
labelDepa.grid(row=5, column=0, padx=10, pady=5)
entryDepa = tk.Entry(Ventana, width=40)
entryDepa.grid(row=5, column=1, padx=10, pady=5)

labelNombreSolicitante = tk.Label(Ventana, text="Nombre del solicitante :")
labelNombreSolicitante.grid(row=6, column=0, padx=10, pady=5)
entryNombreSolicitante = tk.Entry(Ventana, width=40)
entryNombreSolicitante.grid(row=6, column=1, padx=10, pady=5)

labelNumOrden = tk.Label(Ventana, text="Número de orden :")
labelNumOrden.grid(row=7, column=0, padx=10, pady=5)
entryNumOrden = tk.Entry(Ventana, width=40)
entryNumOrden.grid(row=7, column=1, padx=10, pady=5)

labelFecha = tk.Label(Ventana, text="Fecha de la orden :")
labelFecha.grid(row=8, column=0, padx=10, pady=5)
entryFecha = tk.Entry(Ventana, width=40)
entryFecha.grid(row=8, column=1, padx=10, pady=5)

cargar_info_maquina(nombre=maquinaSelected)

# Botón para procesar los datos
btnGenerarPDF = tk.Button(Ventana, text="Generar PDF", command=btnGenerar, bg=colores.AzulMedio, fg="white",width=14, font=("Seoge UI", 10))
btnGenerarPDF.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

btnVolver = tk.Button(Ventana, text="Volver",bg=colores.AzulSucio,fg="white",width=12, font=("Seoge UI", 8), command=volverRecomendaciones)
btnVolver.grid(row=10, column=0, padx=10, pady=5)

btnMenu = tk.Button(Ventana, text="Ir al menú",bg=colores.AzulSucio,fg="white",width=12, font=("Seoge UI", 8), command=volverMenu)
btnMenu.grid(row=10, column=1,  padx=10, pady=5)

# Iniciar la aplicación
Ventana.mainloop()
#cv2.destroyAllWindows()